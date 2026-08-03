import warnings
import openpyxl
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from pathlib import Path

warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="Consejo Ancora 2026",
    page_icon=None,
    layout="wide",
    initial_sidebar_state="expanded",
)

# Rutas relativas al propio archivo: funcionan igual en Windows y en Linux (la nube).
BASE = Path(__file__).parent
EXCEL_PATH = BASE / "EXCEL" / "Presentacion Consejo Junio 2026.xlsx"
LOGO_PATH = BASE / "imagenes" / "ancora_blanco.png"

# ── Paleta de marca Ancora ──────────────────────────────────────────────────────
MARCA_NAVY  = "#13375c"
MARCA_CLARO = "#F6F6F6"
MARCA_TEAL  = "#1d7b8a"
MARCA_ORO   = "#e9ba40"

# Lineas de negocio (colores de marca extendidos).
COLORES_LDN = {
    "BENEFICIOS": MARCA_TEAL,
    "LF":         MARCA_ORO,
    "DAÑOS":      "#3a6ea5",
    "FIANZAS":    "#d97b3c",
    "LP":         "#4aa3b0",
    "AUTOS":      "#8a6d4b",
    "ANI":        "#b58fb0",
    "AFFINITY":   "#9aa5b1",
    "BONO":       "#6f9bc4",
}
COLOR_PTO  = MARCA_ORO   # dorado · meta/presupuesto
COLOR_2026 = MARCA_TEAL  # teal · real (protagonista)
COLOR_2025 = "#3a6ea5"   # navy acero · año anterior
C_NEG = "#e05c5c"
C_POS = MARCA_TEAL

MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# ── Helpers ────────────────────────────────────────────────────────────────────
def fmt(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"${v:,.0f}"

def fmt_m(v):
    # Importe exacto, sin redondear a millones ni a miles (mismo formato que fmt).
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"${v:,.0f}"

def pct(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    return f"{'+' if v >= 0 else ''}{v*100:.1f}%"

def sf(v):
    return float(v) if isinstance(v, (int, float)) else None

def norm(s):
    if not isinstance(s, str):
        return s
    return (s.replace("DA�OS", "DAÑOS").replace("DANOS", "DAÑOS")
             .replace("L�neas", "Líneas").replace("�REA", "AREA")
             .replace("�", "ñ").strip())

# ── Carga ──────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="Cargando datos...")
def load_wb():
    return openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# Conceptos del P&L (General y LdN comparten casi el mismo orden).
CONC_GENERAL = [
    "Base LIne", "Base Line", "Bono", "Venta Nueva", "Ingresos",
    "Costos de Referencia", "Utilidad Bruta", "Sueldos Directos", "Gastos Directos",
    "Utilidad Operativa", "Gastos Indirectos", "Sueldos Corporativos",
    "Sueldos Compartidos", "Utilidad",
]

def parse_general(wb):
    """General: P&L acumulado al 2do trimestre (Ene-Jun), por concepto.
    Columnas: B=PTO 2Q, D=Real 2Q, G=2025 2Q, J=PTO anual, M=2025 anual."""
    ws = wb["General"]
    cols = {"pto_2q": 2, "real_2q": 4, "y25_2q": 7, "pto_anual": 10, "y25_anual": 13}
    data = {}
    for r in range(4, 30):
        c = norm(ws.cell(r, 1).value)
        if not isinstance(c, str) or not c:
            continue
        data[c] = {k: sf(ws.cell(r, col).value) for k, col in cols.items()}
    return data

def parse_ldn(wb):
    """LdN: 7 bloques en dos columnas (A y K), filas 2/29/56/82.
    Dentro de cada bloque: col+1=PTO, col+3=Real, col+6=2025."""
    ws = wb["LdN"]
    bloques = {
        "BENEFICIOS": (2, 1), "LF": (2, 11),
        "FIANZAS": (29, 1), "AUTOS": (29, 11),
        "DAÑOS": (56, 1), "LP": (56, 11),
        "AFFINITY": (82, 1),
    }
    offsets = {"pto": 1, "real": 3, "y25": 6}
    result = {}
    for nombre, (r0, c0) in bloques.items():
        blk = {}
        for r in range(r0 + 1, r0 + 14):
            cc = norm(ws.cell(r, c0).value)
            if not isinstance(cc, str) or not cc or cc == "Concepto":
                continue
            blk[cc] = {k: sf(ws.cell(r, c0 + off).value) for k, off in offsets.items()}
        result[nombre] = blk
    return result

def parse_ingresos(wb):
    """Ingresos: secciones mensuales por area (INGRESOS BRUTOS 2026, VN, PTO)."""
    ws = wb["Ingresos"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    areas = ["AFFINITY", "ANI", "AUTOS", "BENEFICIOS", "DAÑOS",
             "INTERNACIONAL", "FIANZAS", "LF", "LP", "BONO"]

    def val(v):
        return float(v) if isinstance(v, (int, float)) else 0.0

    result = {"2026": {}, "2025": {}, "ppto_2026": {}}
    section = None
    pend_2025 = False   # el bloque de ingresos 2025 no trae rótulo de año:
                        # viene justo después de "COSTOS DE REFERENCIA 2025 REAL".
    for row in rows:
        c0 = norm(row[0]) if row[0] is not None else None
        c1 = norm(row[1]) if len(row) > 1 and row[1] is not None else None
        if c1 == "INGRESOS BRUTOS 2026":
            section = "2026"; continue
        if isinstance(c0, str) and "COSTOS DE REFERENCIA 2025" in c0.upper():
            section = None; pend_2025 = True; continue
        if isinstance(c0, str) and c0.upper() == "AREA" and pend_2025:
            section = "2025"; pend_2025 = False; continue
        if isinstance(c0, str) and "BASE LINE PRESUPUESTO 2026" in c0.upper():
            section = "ppto_2026"; continue
        if c1 in ("INGRESOS BRUTOS 2026 VENTA NUEVA", "VENTA NUEVA 2025",
                  "VENTA NUEVA PRESUPUESTO 2026") or \
           (isinstance(c1, str) and "VENTA N" in c1.upper()):
            section = None; continue
        if isinstance(c0, str) and c0.upper() in [a.upper() for a in areas] and section:
            vals = [val(row[i]) for i in range(1, 13)]
            key = c0.upper()
            if key not in result[section]:
                result[section][key] = vals
            else:
                result[section][key] = [a + b for a, b in zip(result[section][key], vals)]
    return result

try:
    wb = load_wb()
    GEN = parse_general(wb)
    LDN = parse_ldn(wb)
    ING = parse_ingresos(wb)
except Exception as e:
    st.error(f"Error al abrir el archivo:\n\n`{EXCEL_PATH}`\n\n{e}")
    st.stop()

def g(concepto, campo):
    return GEN.get(concepto, {}).get(campo)

# Ingresos usa 'Base LIne' (con el typo del Excel) en General.
BASE_KEY = "Base LIne" if "Base LIne" in GEN else "Base Line"

# AFFINITY se excluye (viene en $0 en el archivo de junio).
LDN_LIST = ["BENEFICIOS", "LF", "FIANZAS", "AUTOS", "DAÑOS", "LP"]

# ── Estilos (marca Ancora) ──────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700&display=swap');
html, body, [data-testid="stAppViewContainer"], [data-testid="stSidebar"], .stApp,
.stApp *:not([data-testid="stIconMaterial"]) { font-family: 'Montserrat', sans-serif !important; }
[data-testid="stIconMaterial"] { font-family: 'Material Symbols Rounded' !important; }
[data-testid="stAppViewContainer"], [data-testid="stMarkdownContainer"],
p, span:not([data-testid="stIconMaterial"]), label, td, th, li,
div[data-testid="stMetricValue"], div[data-testid="stMetricDelta"] { font-weight: 500; }
h1, h2, h3, h4, h5, h6, div[data-testid="stMetricLabel"] p { font-weight: 600 !important; }
section[data-testid="stSidebar"] img { pointer-events: none; }
section[data-testid="stSidebar"] [data-testid="StyledFullScreenButton"] { display: none; }
section[data-testid="stSidebar"] > div:first-child { padding-top: 0 !important; }
div[data-testid="stRadio"] > div { gap: 2px !important; }
div[data-testid="stRadio"] label {
    display: flex !important; align-items: center !important;
    padding: 10px 14px !important; border-radius: 8px !important;
    cursor: pointer !important; font-size: 0.9rem !important;
    margin-bottom: 2px !important; transition: background 0.15s !important; border: none !important;
}
div[data-testid="stRadio"] label:hover { background: rgba(255,255,255,0.08) !important; }
div[data-testid="stRadio"] label:has(input:checked) {
    background: rgba(255,255,255,0.13) !important; font-weight: 600 !important;
}
div[data-testid="stRadio"] label > div:first-child { display: none !important; }
div[data-testid="stRadio"] { border: none !important; }
</style>
""", unsafe_allow_html=True)

# Evita que el traductor del navegador reformatee los montos.
components.html(
    "<script>try{var d=window.parent.document.documentElement;"
    "d.setAttribute('translate','no');d.classList.add('notranslate');}catch(e){}</script>",
    height=0,
)

with st.sidebar:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), use_container_width=True)
    st.markdown("---")
    st.caption("TABLEROS")
    pagina = st.radio(
        "",
        ["Resumen", "Por Línea de Negocio"],
        label_visibility="collapsed",
        key="nav",
    )
    st.markdown("---")
    st.caption("Consejo Ancora · Cierre 2T 2026 (Ene–Jun)")

# HTML de tabla P&L (tema navy de marca).
PL_CSS = f"""
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Montserrat',-apple-system,'Segoe UI',sans-serif;background:transparent;
color:{MARCA_CLARO};font-size:13px;font-weight:500}}
table{{width:100%;border-collapse:collapse}}
th{{padding:8px 12px;text-align:right;border-bottom:2px solid #2e4d70;color:#9db3cc;
font-weight:600;font-size:12px;white-space:nowrap}}
th:first-child{{text-align:left}}
td{{padding:7px 12px;text-align:right;border-bottom:1px solid #21344c;white-space:nowrap}}
td:first-child{{text-align:left}}
.main td{{font-weight:700}}
.sub{{display:none}}
.sub td{{font-size:12px;color:#a9bad0;font-weight:400}}
.sc{{padding-left:30px!important}}
.clk{{cursor:pointer}}
.clk:hover td{{background:rgba(255,255,255,.06)}}
.sub:hover td{{background:rgba(255,255,255,.03)}}
.neg{{color:{C_NEG}}}.pos{{color:{MARCA_TEAL}}}
.arr{{font-size:9px;color:#9db3cc;display:inline-block;width:12px;text-align:center}}
"""
PL_JS = """
function tgl(row){
  var gid=row.getAttribute('data-g');
  var subs=document.querySelectorAll('.sub-'+gid);
  var arr=document.getElementById('a-'+gid);
  var open=arr.textContent==='▼';
  subs.forEach(function(s){s.style.display=open?'none':'table-row'});
  arr.textContent=open?'►':'▼';
}
"""

st.title("Tablero Ejecutivo Ancora · 2T 2026")
st.caption("Acumulado al segundo trimestre (Ene–Jun) · Importes en MXN")

# ══════════════════════════════════════════════════════════════════════════════
# RESUMEN
# ══════════════════════════════════════════════════════════════════════════════
if pagina == "Resumen":
    ingreso_real = g("Ingresos", "real_2q")
    kpis = [
        ("Ingresos", "Ingresos"),
        ("Utilidad Bruta", "Utilidad Bruta"),
        ("Utilidad Operativa", "Utilidad Operativa"),
        ("EBITDA", "Utilidad"),
    ]
    cols = st.columns(4)
    for col, (label, key) in zip(cols, kpis):
        real, ptov, y25 = g(key, "real_2q"), g(key, "pto_2q"), g(key, "y25_2q")
        d25 = (real / y25 - 1) if real and y25 else None
        margen = (real / ingreso_real) if (real and ingreso_real and key != "Ingresos") else None
        avance = (real / ptov) if (real and ptov) else None
        with col:
            st.metric(label, fmt_m(real),
                      delta=(f"Margen {pct(margen)}" if margen is not None else None),
                      delta_color="off")
            st.caption(f"PTO: {fmt_m(ptov)}"
                       + (f" · avance {avance*100:.1f}%" if avance is not None else "")
                       + f"  |  2025: {fmt_m(y25)}  ({pct(d25)} vs 2025)")

    st.divider()
    st.subheader("PTO vs Real vs 2025 — Indicadores clave")
    keys = ["Ingresos", "Utilidad Bruta", "Utilidad Operativa", "Utilidad"]
    labels = ["Ingresos", "Util. Bruta", "Util. Operativa", "EBITDA"]
    fig = go.Figure()
    for name, field, color in [("PTO 2Q", "pto_2q", COLOR_PTO),
                               ("Real 2Q", "real_2q", COLOR_2026),
                               ("2025 2Q", "y25_2q", COLOR_2025)]:
        vals = [g(k, field) or 0 for k in keys]
        ing = g("Ingresos", field) or 0
        # Cifra exacta + % de margen (de cada utilidad sobre ingresos).
        txt = [fmt(v) if i == 0 else f"{fmt(v)}<br>{pct(v/ing) if ing else '—'}"
               for i, v in enumerate(vals)]
        fig.add_bar(name=name, x=labels, y=vals, marker_color=color,
                    text=txt, textposition="outside",
                    textfont=dict(size=10, color="#e0e0e0"), cliponaxis=False)
    # % de crecimiento vs 2025 en medio de la barra Real 2Q (barra central del grupo).
    for lab, k in zip(labels, keys):
        r, q = g(k, "real_2q"), g(k, "y25_2q")
        if r and q:
            fig.add_annotation(x=lab, y=r / 2, text=pct(r / q - 1), showarrow=False,
                               font=dict(size=13, color="#ffffff"),
                               bgcolor="rgba(15,33,55,0.6)", borderpad=3)
    # Sobre la barra amarilla (PTO): línea al nivel del Real = % de avance (Real/PTO).
    _grp_w, _nbar = 0.8, 3
    _bar_w = _grp_w / _nbar
    _pto_off = -_grp_w / 2 + _bar_w / 2   # 1a barra del grupo (PTO), respecto al centro
    for i, k in enumerate(keys):
        p, r = g(k, "pto_2q"), g(k, "real_2q")
        if p and r:
            fig.add_shape(type="line", xref="x", yref="y",
                          x0=i + _pto_off - _bar_w / 2, x1=i + _pto_off + _bar_w / 2,
                          y0=r, y1=r, line=dict(color="#0f2137", width=3))
            fig.add_annotation(x=i + _pto_off, y=r, xref="x", yref="y",
                               text=f"avance {r / p * 100:.1f}%", showarrow=False,
                               yanchor="bottom", yshift=3,
                               font=dict(size=10, color="#0f2137"),
                               bgcolor="rgba(255,255,255,0.8)", borderpad=2)
    fig.update_layout(dragmode="pan", barmode="group", yaxis_tickformat="$,.0f",
                      legend=dict(orientation="h", y=1.12),
                      uniformtext=dict(minsize=7, mode="hide"),
                      height=500, margin=dict(t=80, b=20))
    st.plotly_chart(fig, use_container_width=True)

    st.divider()
    st.subheader("Estado de Resultados · Acumulado 2T 2026")

    # Estructura idéntica al Excel (hoja General, filas 4–16), hasta EBITDA:
    # Base/Bono/VN → Ingresos → Costos → Utilidad Bruta → Sueldos/Gastos Dir →
    # Utilidad Operativa → Gastos Ind/Sueldos Corp/Comp → EBITDA.
    grupos = [
        ("ing", "Ingresos Totales", "Ingresos", [
            ("Base Line", BASE_KEY), ("Bono", "Bono"), ("Venta Nueva", "Venta Nueva")]),
        (None, "Costos de Referencia", "Costos de Referencia", []),
        (None, "Utilidad Bruta", "Utilidad Bruta", []),
        ("uo", "Utilidad Operativa", "Utilidad Operativa", [
            ("Sueldos Directos", "Sueldos Directos"), ("Gastos Directos", "Gastos Directos")]),
        ("eb", "EBITDA", "Utilidad", [
            ("Gastos Indirectos", "Gastos Indirectos"),
            ("Sueldos Corporativos", "Sueldos Corporativos"),
            ("Sueldos Compartidos", "Sueldos Compartidos")]),
    ]

    # Ingresos de referencia (por columna) para el % de integración.
    ing_p = g("Ingresos", "pto_2q") or 0
    ing_r = g("Ingresos", "real_2q") or 0
    ing_q = g("Ingresos", "y25_2q") or 0

    def _int(v, base):
        return f"{v/base*100:.1f}%" if (v is not None and base) else "—"

    def rv(key):
        r, p, q = g(key, "real_2q"), g(key, "pto_2q"), g(key, "y25_2q")
        return (fmt(p), _int(p, ing_p), fmt(r), _int(r, ing_r),
                pct((r / p - 1) if r and p else None),
                fmt(q), _int(q, ing_q),
                pct((r / q - 1) if r and q else None))

    def cc(v):
        if v == "—":
            return ""
        return "neg" if v.startswith("-") else "pos"

    def fila(label, key, gid=None, sub=False):
        pv, pi, rvv, ri, vp, qv, qi, v25 = rv(key)
        cls = f"sub sub-{gid}" if sub else ("main clk" if gid else "main")
        clk = f' data-g="{gid}" onclick="tgl(this)"' if (gid and not sub) else ""
        first = (f'<span class="arr" id="a-{gid}">►</span> {label}' if (gid and not sub)
                 else label)
        c1 = "sc" if sub else ""
        return (f'<tr class="{cls}"{clk}><td class="{c1}">{first}</td>'
                f'<td>{pv}</td><td class="pi">{pi}</td>'
                f'<td>{rvv}</td><td class="pi">{ri}</td>'
                f'<td class="{cc(vp)}">{vp}</td>'
                f'<td>{qv}</td><td class="pi">{qi}</td>'
                f'<td class="{cc(v25)}">{v25}</td></tr>')

    body = ""
    for gid, label, key, subs in grupos:
        for sl, sk in subs:
            body += fila(sl, sk, gid=gid, sub=True)
        body += fila(label, key, gid=gid)

    html = (f'<!DOCTYPE html><html><head><meta charset="utf-8"><style>{PL_CSS}'
            f'.pi{{color:#9db3cc;font-size:11px}}</style></head>'
            f'<body><table><thead><tr>'
            f'<th>Concepto</th><th>PTO 2Q</th><th>% Int</th>'
            f'<th>Real 2Q</th><th>% Int</th><th>vs PTO</th>'
            f'<th>2025 2Q</th><th>% Int</th><th>vs 2025</th></tr></thead>'
            f'<tbody>{body}</tbody></table><script>{PL_JS}</script></body></html>')
    components.html(html, height=430, scrolling=False)
    st.caption("Da clic en un renglón principal para ver su desglose. "
               "% Int = cada concepto como porcentaje de los ingresos.")

# ══════════════════════════════════════════════════════════════════════════════
# POR LÍNEA DE NEGOCIO
# ══════════════════════════════════════════════════════════════════════════════
if pagina == "Por Línea de Negocio":
    st.subheader("Comparativo por Línea de Negocio · 2T 2026")

    def lv(ldn, concepto, campo):
        return LDN.get(ldn, {}).get(concepto, {}).get(campo) or 0

    total_ing = sum(lv(l, "Ingresos", "real") for l in LDN_LIST)
    # Orden descendente por ingreso real (2T).
    ldn_orden = sorted(LDN_LIST, key=lambda l: -lv(l, "Ingresos", "real"))
    # Carrusel: 3 tarjetas a la vez, con flechas para ver las demas.
    PER_PAGE = 3
    total_pages = -(-len(ldn_orden) // PER_PAGE)  # division techo
    if "ldn_page" not in st.session_state:
        st.session_state.ldn_page = 0
    page = max(0, min(st.session_state.ldn_page, total_pages - 1))
    fila = ldn_orden[page * PER_PAGE:(page + 1) * PER_PAGE]

    cols = st.columns(3)
    for col, l in zip(cols, fila):
        real = lv(l, "Ingresos", "real")
        ptov = lv(l, "Ingresos", "pto")
        y25 = lv(l, "Ingresos", "y25")
        d25 = (real / y25 - 1) if real and y25 else None
        margen = (real / total_ing) if real and total_ing else None
        with col:
            st.metric(l, fmt_m(real),
                      delta=(f"{margen*100:.1f}% del total" if margen is not None else None),
                      delta_color="off")
            st.caption(f"PTO {fmt_m(ptov)} · {pct(d25)} vs 25")

    c_prev, c_ind, c_next = st.columns([1, 4, 1])
    with c_prev:
        if st.button("← Anterior", disabled=(page == 0), use_container_width=True):
            st.session_state.ldn_page = page - 1
            st.rerun()
    with c_ind:
        st.markdown(f"<div style='text-align:center;color:#9db3cc;font-size:13px;"
                    f"padding-top:8px'>Página {page + 1} de {total_pages}</div>",
                    unsafe_allow_html=True)
    with c_next:
        if st.button("Siguiente →", disabled=(page >= total_pages - 1),
                     use_container_width=True):
            st.session_state.ldn_page = page + 1
            st.rerun()

    st.divider()
    st.subheader("Ingreso por área · primer semestre: 2025, 2026 y presupuesto")
    _areas_i = ["BENEFICIOS", "FIANZAS", "DAÑOS", "LF", "AUTOS", "LP", "ANI", "BONO"]

    def _ser(bloque, a):
        return ING.get(bloque, {}).get(a, [0] * 12)

    _nm = 0
    for m in range(12):
        if sum(_ser("2026", a)[m] for a in _areas_i) > 0:
            _nm = m + 1
    r26 = {a: sum(_ser("2026", a)[:_nm]) for a in _areas_i}
    r25 = {a: sum(_ser("2025", a)[:_nm]) for a in _areas_i}
    pp = {a: sum(_ser("ppto_2026", a)[:_nm]) for a in _areas_i}
    ao = sorted([a for a in _areas_i if r26.get(a, 0) or r25.get(a, 0)],
                key=lambda a: -r26.get(a, 0))
    # 2026: crecimiento vs 2025. PTO: avance (Real 2026 ÷ PTO del semestre).
    crec = [f"{r26[a]/r25[a]*100-100:+.0f}%" if r25.get(a) else "" for a in ao]
    avance = [f"av. {r26[a]/pp[a]*100:.0f}%" if pp.get(a) else "" for a in ao]
    fig_sem = go.Figure()
    fig_sem.add_bar(name="1er sem 2025", x=ao, y=[r25.get(a, 0) for a in ao],
                    marker_color=COLOR_2025)
    fig_sem.add_bar(name="1er sem 2026", x=ao, y=[r26.get(a, 0) for a in ao],
                    marker_color=COLOR_2026, text=crec, textposition="outside",
                    textfont=dict(size=10, color="#e0e0e0"), cliponaxis=False)
    fig_sem.add_bar(name="PTO 2026", x=ao, y=[pp.get(a, 0) for a in ao],
                    marker_color=COLOR_PTO, text=avance, textposition="outside",
                    textfont=dict(size=10, color="#e0e0e0"), cliponaxis=False)
    fig_sem.update_layout(dragmode="pan", barmode="group", yaxis_tickformat="$,.0f", height=470,
                          margin=dict(t=50, b=20), legend=dict(orientation="h", y=1.1))
    st.plotly_chart(fig_sem, use_container_width=True)

    st.divider()
    st.subheader("Ingresos (barras) y EBITDA (línea) por LdN")
    st.caption("Los 3 periodos a la vez · barras = Ingresos, líneas = EBITDA · "
               "2025, Real 2026 y PTO, en la misma escala")
    fig = go.Figure()
    for etq, fld, color in [("Ing. 2025", "y25", COLOR_2025),
                            ("Ing. Real 2026", "real", COLOR_2026),
                            ("Ing. PTO", "pto", COLOR_PTO)]:
        fig.add_bar(name=etq, x=LDN_LIST,
                    y=[lv(l, "Ingresos", fld) for l in LDN_LIST], marker_color=color)
    # Colores de línea distintos a los de las barras para que contrasten.
    for etq, fld, color, dash, w, tpos in [
            ("EBITDA 2025", "y25", "#aab7c4", "dot", 2, "bottom center"),
            ("EBITDA Real 2026", "real", "#ff7a5c", None, 4, "top center"),
            ("EBITDA PTO", "pto", "#f2f2f2", "dash", 2.5, "top center")]:
        # En cada punto: % = EBITDA / Ingresos de esa LdN (margen EBITDA).
        rent_txt = [f"{lv(l, 'EBITDA', fld) / lv(l, 'Ingresos', fld) * 100:.1f}%"
                    if lv(l, "Ingresos", fld) else "" for l in LDN_LIST]
        fig.add_scatter(name=etq, x=LDN_LIST,
                        y=[lv(l, "EBITDA", fld) for l in LDN_LIST],
                        mode="lines+markers+text",
                        text=rent_txt, textposition=tpos,
                        textfont=dict(size=9, color=color),
                        line=dict(color=color, width=w, dash=dash),
                        marker=dict(size=7, line=dict(width=1, color="#0f2137")))
    fig.update_layout(dragmode="pan", barmode="group", height=520, margin=dict(t=50, b=10),
                      legend=dict(orientation="h", y=1.14),
                      yaxis=dict(title="MXN", tickformat="$,.0f"))
    st.plotly_chart(fig, use_container_width=True)

    st.divider()
    # Selector: controla el Mix y la tabla de abajo.
    YEAR_OPTS = {"Real 2026": "real", "PTO 2026": "pto", "2025": "y25"}
    year_sel = st.radio("Año para el Mix y la tabla",
                        list(YEAR_OPTS.keys()), horizontal=True, key="ldn_year")
    field = YEAR_OPTS[year_sel]
    cp, ct = st.columns([2, 3])
    with cp:
        st.subheader(f"Mix de Ingresos · {year_sel}")
        vals = {l: lv(l, "Ingresos", field) for l in LDN_LIST if lv(l, "Ingresos", field) > 0}
        fig_p = go.Figure(go.Pie(
            labels=list(vals.keys()), values=list(vals.values()),
            marker_colors=[COLORES_LDN.get(l, "#999") for l in vals],
            hole=0.42, textinfo="label+percent", sort=True))
        fig_p.update_layout(height=340, margin=dict(t=20, b=10), showlegend=False)
        st.plotly_chart(fig_p, use_container_width=True)
    with ct:
        st.subheader(f"Tabla resumen por LdN · {year_sel}")
        # (clave en los datos, etiqueta a mostrar). EBITDA vive como "Utilidad".
        conceptos = [("Ingresos", "Ingresos"),
                     ("Utilidad Bruta", "Utilidad Bruta"),
                     ("Utilidad Operativa", "Utilidad Operativa"),
                     ("EBITDA", "EBITDA")]
        # Total por concepto (del año seleccionado) para el % de integración.
        totales = {c: sum(lv(l, c, field) for l in LDN_LIST) for c, _ in conceptos}
        rows = []
        for l in LDN_LIST:
            row = {"LdN": l}
            for c, lab in conceptos:
                val = lv(l, c, field)
                ptov = lv(l, c, "pto")
                row[lab] = fmt(val)
                # % de integración: cuánto representa esta LdN del total del concepto.
                row[f"{lab} % Int"] = f"{val/totales[c]*100:.1f}%" if totales[c] else "—"
                # vs PTO solo aplica al Real 2026 (no se compara 2025 ni el PTO contra sí mismo).
                row[f"{lab} vs PTO"] = pct((val / ptov - 1) if ptov else None) if field == "real" else "—"
            # Rentabilidad = EBITDA ÷ Ingresos de esa línea (margen EBITDA).
            ing_l = lv(l, "Ingresos", field)
            row["Rentabilidad"] = f"{lv(l, 'EBITDA', field)/ing_l*100:.1f}%" if ing_l else "—"
            rows.append(row)
        # LdN como índice: st.dataframe la deja fija al hacer scroll horizontal.
        df_ldn = pd.DataFrame(rows).set_index("LdN")
        int_cols = [c for c in df_ldn.columns if c.endswith("% Int")]
        sty = df_ldn.style.set_properties(
            subset=int_cols,
            **{"background-color": "rgba(31,119,180,0.30)", "color": "#dbeeff"})
        st.dataframe(sty, use_container_width=True)

    st.divider()
    st.subheader("Detalle mensual")
    areas_ing = ["BENEFICIOS", "FIANZAS", "DAÑOS", "LF", "AUTOS", "LP", "ANI", "BONO"]

    def _serie(a):
        return ING.get("2026", {}).get(a, [0] * 12)

    n_mes = 0
    for m in range(12):
        if sum(_serie(a)[m] for a in areas_ing) > 0:
            n_mes = m + 1
    filas = []
    for a in areas_ing:
        s = _serie(a)
        if sum(s) == 0:
            continue
        fila = {"Área": a}
        for m in range(n_mes):
            fila[MESES[m]] = fmt(s[m])
        fila["Total"] = fmt(sum(s[:n_mes]))
        filas.append(fila)
    tot = {"Área": "TOTAL"}
    for m in range(n_mes):
        tot[MESES[m]] = fmt(sum(_serie(a)[m] for a in areas_ing))
    tot["Total"] = fmt(sum(sum(_serie(a)[:n_mes]) for a in areas_ing))
    filas.append(tot)
    st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("Detalle por Línea de Negocio")
    sel = st.selectbox("Selecciona una LdN", LDN_LIST)
    d_sel = LDN.get(sel, {})
    grupos = [
        ("ing", "Ingresos", "Ingresos", [
            ("Base Line", "Base Line"),
            # El rubro Bono aparece como "Bono" en unas LdN y "Bonos" en otras.
            ("Bono", "Bono"), ("Bono", "Bonos"),
            ("Venta Nueva", "Venta Nueva")]),
        (None, "Costos de Referencia", "Costos de Referencia", []),
        (None, "Utilidad Bruta", "Utilidad Bruta", []),
        ("uo", "Utilidad Operativa", "Utilidad Operativa", [
            ("Sueldos Directos", "Sueldos Directos"), ("Gastos Directos", "Gastos Directos")]),
        ("eb", "EBITDA", "EBITDA", [
            ("Gastos Indirectos", "Gastos Indirectos"),
            ("Sueldos Corporativos", "Sueldos Corporativos"),
            ("Sueldos Compartidos", "Sueldos Compartidos")]),
    ]

    # Ingresos de la LdN (por columna) para el % de integración.
    ing = d_sel.get("Ingresos", {})
    ip, ir, iq = ing.get("pto") or 0, ing.get("real") or 0, ing.get("y25") or 0

    def _int(v, base):
        return f"{v/base*100:.1f}%" if (v is not None and base) else "—"

    def dv(key):
        d = d_sel.get(key, {})
        r, p, q = d.get("real"), d.get("pto"), d.get("y25")
        return (fmt(p), _int(p, ip), fmt(r), _int(r, ir),
                pct((r / p - 1) if r and p else None),
                fmt(q), _int(q, iq),
                pct((r / q - 1) if r and q else None))

    def cc2(v):
        if v == "—":
            return ""
        return "neg" if v.startswith("-") else "pos"

    def fila2(label, key, gid=None, sub=False):
        pv, pi, rvv, ri, vp, qv, qi, v25 = dv(key)
        cls = f"sub sub-{gid}" if sub else ("main clk" if gid else "main")
        clk = f' data-g="{gid}" onclick="tgl(this)"' if (gid and not sub) else ""
        first = (f'<span class="arr" id="a-{gid}">►</span> {label}' if (gid and not sub)
                 else label)
        c1 = "sc" if sub else ""
        return (f'<tr class="{cls}"{clk}><td class="{c1}">{first}</td>'
                f'<td>{pv}</td><td class="pi">{pi}</td>'
                f'<td>{rvv}</td><td class="pi">{ri}</td>'
                f'<td class="{cc2(vp)}">{vp}</td>'
                f'<td>{qv}</td><td class="pi">{qi}</td>'
                f'<td class="{cc2(v25)}">{v25}</td></tr>')

    body = ""
    for gid, label, key, subs in grupos:
        for sl, sk in subs:
            if sk not in d_sel:
                continue
            body += fila2(sl, sk, gid=gid, sub=True)
        body += fila2(label, key, gid=gid)

    html = (f'<!DOCTYPE html><html><head><meta charset="utf-8"><style>{PL_CSS}'
            f'.pi{{color:#9db3cc;font-size:11px}}</style></head>'
            f'<body><table><thead><tr>'
            f'<th>Concepto</th><th>PTO 2026</th><th>% Int</th>'
            f'<th>Real 2T</th><th>% Int</th><th>vs PTO</th>'
            f'<th>2025</th><th>% Int</th><th>vs 2025</th></tr></thead>'
            f'<tbody>{body}</tbody></table><script>{PL_JS}</script></body></html>')
    components.html(html, height=420, scrolling=False)

